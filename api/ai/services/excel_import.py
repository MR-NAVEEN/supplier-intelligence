import csv
import json
import os
import re
from io import StringIO

from openpyxl import load_workbook

from api.ai.services.extraction import _clean_price_value

CANONICAL_FIELDS = (
    'product_name',
    'code_or_sku',
    'price',
    'currency',
    'description',
    'series',
    'page_number',
)

HEADER_SYNONYMS = {
    'product_name': {
        'product name',
        'product',
        'item',
        'item name',
        'name',
        'model',
        'model name',
        'goods',
        'description of goods',
        'product title',
    },
    'code_or_sku': {
        'sku',
        'code',
        'item code',
        'product code',
        'model no',
        'model number',
        'model no.',
        'article',
        'article no',
        'hsn',
        'part no',
        'part number',
        'item no',
    },
    'price': {
        'price',
        'mrp',
        'rate',
        'amount',
        'unit price',
        'selling price',
        'dp',
        'dealer price',
        'price (mrp)',
        'mrp price',
    },
    'currency': {'currency', 'curr', 'ccy'},
    'description': {'description', 'details', 'spec', 'specification', 'remarks'},
    'series': {'series', 'category', 'brand', 'range', 'collection'},
    'page_number': {'page', 'page no', 'page number', 'pg'},
}

MAX_HEADER_SCAN_ROWS = 20
MAX_EMPTY_STREAK = 25


def _norm_header(value):
    text = str(value or '').strip().lower()
    text = text.replace('\n', ' ')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def guess_field(header):
    key = _norm_header(header)
    if not key:
        return None
    best = None
    best_len = -1
    for field, aliases in HEADER_SYNONYMS.items():
        candidates = set(aliases)
        candidates.add(field.replace('_', ' '))
        for alias in candidates:
            if key == alias and len(alias) > best_len:
                best = field
                best_len = len(alias)
    return best


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def detect_header_row(rows):
    best_index = 0
    best_score = -1
    limit = min(len(rows), MAX_HEADER_SCAN_ROWS)
    for index in range(limit):
        mapped = [guess_field(cell) for cell in rows[index]]
        score = sum(1 for field in mapped if field)
        filled = sum(1 for cell in rows[index] if _cell_text(cell))
        if score > best_score and filled >= 2:
            best_score = score
            best_index = index
    return best_index, best_score


def build_column_map(headers, user_map=None):
    """Map canonical field -> column index."""
    mapping = {}
    user_map = user_map or {}
    normalized_headers = [_norm_header(h) for h in headers]
    for field, header_name in user_map.items():
        if field not in CANONICAL_FIELDS:
            continue
        wanted = _norm_header(header_name)
        if wanted in normalized_headers:
            mapping[field] = normalized_headers.index(wanted)
        else:
            try:
                mapping[field] = int(header_name)
            except (TypeError, ValueError):
                continue
    for index, header in enumerate(headers):
        field = guess_field(header)
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def parse_user_column_map(raw):
    if not raw:
        return {}
    if isinstance(raw, dict):
        return {str(k): str(v) for k, v in raw.items() if str(k) in CANONICAL_FIELDS}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError('column_map must be valid JSON, e.g. {"product_name":"Item","price":"MRP"}.') from exc
    if not isinstance(data, dict):
        raise ValueError('column_map must be a JSON object.')
    return {str(k): str(v) for k, v in data.items() if str(k) in CANONICAL_FIELDS}


def _read_csv_rows(path):
    with open(path, 'rb') as handle:
        raw = handle.read()
    text = raw.decode('utf-8-sig', errors='replace')
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(text), dialect)
    return [list(row) for row in reader]


def _read_xlsx_rows(path, sheet=None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet in (None, ''):
            worksheet = workbook.worksheets[0]
        else:
            try:
                worksheet = workbook.worksheets[int(sheet)]
            except (TypeError, ValueError, IndexError):
                if sheet not in workbook.sheetnames:
                    raise ValueError(
                        f'Sheet {sheet!r} not found. Available: {", ".join(workbook.sheetnames)}'
                    )
                worksheet = workbook[sheet]
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([cell for cell in row])
        return rows, worksheet.title
    finally:
        workbook.close()


def rows_to_products(rows, column_map=None, header_row=None, max_rows=5000):
    if not rows:
        raise ValueError('The spreadsheet is empty.')
    if header_row is None:
        header_index, _score = detect_header_row(rows)
    else:
        header_index = max(int(header_row) - 1, 0)
        if header_index >= len(rows):
            raise ValueError('header_row is past the end of the sheet.')
    headers = [_cell_text(cell) or f'column_{i + 1}' for i, cell in enumerate(rows[header_index])]
    mapping = build_column_map(headers, column_map)
    if 'product_name' not in mapping and 'code_or_sku' not in mapping:
        raise ValueError(
            'Could not detect product name or SKU column. '
            'Send column_map, e.g. {"product_name":"Item Name","price":"MRP"}.'
        )

    products = []
    skipped = 0
    empty_streak = 0
    for row in rows[header_index + 1 :]:
        if len(products) >= max_rows:
            break
        values = [_cell_text(cell) for cell in row]
        if not any(values):
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_STREAK:
                break
            skipped += 1
            continue
        empty_streak = 0
        product = {}
        extras = {}
        used = set(mapping.values())
        for field, index in mapping.items():
            if index >= len(values):
                continue
            product[field] = values[index]
        for index, header in enumerate(headers):
            if index in used or index >= len(values):
                continue
            text = values[index]
            if text:
                extras[header] = text
        name = (product.get('product_name') or '').strip()
        sku = (product.get('code_or_sku') or '').strip()
        price = _clean_price_value(product.get('price'))
        if not name and not sku and not price:
            skipped += 1
            continue
        if not name:
            name = sku or extras.get(headers[0], 'Untitled')
        page_raw = product.get('page_number') or '1'
        try:
            page_number = max(int(float(str(page_raw).strip() or 1)), 1)
        except (TypeError, ValueError):
            page_number = 1
        products.append(
            {
                'product_name': name,
                'code_or_sku': sku or None,
                'price': price,
                'currency': (product.get('currency') or '').strip() or None,
                'description': (product.get('description') or '').strip() or None,
                'series': (product.get('series') or '').strip() or None,
                'attributes': extras,
                'page_number': page_number,
            }
        )
    return {
        'headers': headers,
        'header_row': header_index + 1,
        'column_map': {field: headers[index] for field, index in mapping.items() if index < len(headers)},
        'products': products,
        'rows_skipped': skipped,
    }


def import_spreadsheet(path, sheet=None, column_map=None, header_row=None, max_rows=5000):
    ext = os.path.splitext(path)[1].lower()
    sheet_name = None
    if ext in {'.csv', '.tsv'}:
        rows = _read_csv_rows(path)
        sheet_name = os.path.basename(path)
    elif ext in {'.xlsx', '.xlsm'}:
        rows, sheet_name = _read_xlsx_rows(path, sheet=sheet)
    elif ext == '.xls':
        raise ValueError('Legacy .xls is not supported. Save as .xlsx or CSV and upload again.')
    else:
        raise ValueError('Upload a .xlsx, .xlsm, .csv, or .tsv file.')

    parsed = rows_to_products(rows, column_map=column_map, header_row=header_row, max_rows=max_rows)
    by_page = {}
    for product in parsed['products']:
        page_number = int(product.get('page_number') or 1)
        by_page.setdefault(page_number, []).append(product)
    pages = []
    for page_number in sorted(by_page):
        items = by_page[page_number]
        pages.append(
            {
                'page_number': page_number,
                'page_type': 'product_listing',
                'series_or_section_title': items[0].get('series'),
                'products': items,
                'raw_text_summary': f'{len(items)} rows from spreadsheet',
                'page_notes': f'sheet={sheet_name}',
            }
        )
    return {
        'result': {
            'source_file': os.path.basename(path),
            'source_type': 'spreadsheet',
            'sheet': sheet_name,
            'header_row': parsed['header_row'],
            'column_map': parsed['column_map'],
            'total_pages_in_pdf': len(pages) or 1,
            'pages_processed': [page['page_number'] for page in pages],
            'pages': pages,
            'rows_imported': len(parsed['products']),
            'rows_skipped': parsed['rows_skipped'],
        },
        'usage': {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0},
        'advertisement_pages_skipped': 0,
    }
