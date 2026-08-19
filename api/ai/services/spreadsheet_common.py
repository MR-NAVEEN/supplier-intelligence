"""Shared CSV/XLSX reading and header-mapping helpers.

Used by both the product importer (excel_import.py) and the business-card
importer (card_import.py): reading rows, detecting the header row, and
mapping messy header text to a fixed set of canonical field names — first
via a synonym dictionary (instant, free, deterministic), then via an AI
fallback when the dictionary can't confidently map the required fields.
"""
import csv
import json
import os
from io import StringIO

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

from openpyxl import load_workbook

MAX_HEADER_SCAN_ROWS = 20
MAX_EMPTY_STREAK = 25
AI_SAMPLE_ROWS = 3


def _norm_header(value):
    import re

    text = str(value or '').strip().lower()
    text = text.replace('\n', ' ')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def guess_field(header, header_synonyms):
    key = _norm_header(header)
    if not key:
        return None
    best = None
    best_len = -1
    for field, aliases in header_synonyms.items():
        candidates = set(aliases)
        candidates.add(field.replace('_', ' '))
        for alias in candidates:
            if key == alias and len(alias) > best_len:
                best = field
                best_len = len(alias)
    return best


def detect_header_row(rows, header_synonyms):
    best_index = 0
    best_score = -1
    limit = min(len(rows), MAX_HEADER_SCAN_ROWS)
    for index in range(limit):
        mapped = [guess_field(cell, header_synonyms) for cell in rows[index]]
        score = sum(1 for field in mapped if field)
        filled = sum(1 for cell in rows[index] if _cell_text(cell))
        if score > best_score and filled >= 2:
            best_score = score
            best_index = index
    return best_index, best_score


def build_column_map(headers, header_synonyms, canonical_fields, user_map=None):
    """Map canonical field -> column index, using an explicit user_map first, then synonyms."""
    mapping = {}
    user_map = user_map or {}
    normalized_headers = [_norm_header(h) for h in headers]
    for field, header_name in user_map.items():
        if field not in canonical_fields:
            continue
        wanted = _norm_header(header_name)
        if wanted in normalized_headers:
            mapping[field] = normalized_headers.index(wanted)
        else:
            try:
                mapping[field] = int(header_name)
            except (TypeError, ValueError):
                continue
    for index, header in enumerate(headers):
        field = guess_field(header, header_synonyms)
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def parse_user_column_map(raw, canonical_fields):
    if not raw:
        return {}
    if isinstance(raw, dict):
        return {str(k): str(v) for k, v in raw.items() if str(k) in canonical_fields}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError('column_map must be valid JSON, e.g. {"product_name":"Item","price":"MRP"}.') from exc
    if not isinstance(data, dict):
        raise ValueError('column_map must be a JSON object.')
    return {str(k): str(v) for k, v in data.items() if str(k) in canonical_fields}


def read_csv_rows(path):
    with open(path, 'rb') as handle:
        raw = handle.read()
    text = raw.decode('utf-8-sig', errors='replace')
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(text), dialect)
    return [list(row) for row in reader]


def read_xlsx_rows(path, sheet=None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet in (None, ''):
            worksheet = workbook.worksheets[0]
        else:
            try:
                worksheet = workbook.worksheets[int(sheet)]
            except (TypeError, ValueError, IndexError):
                if sheet not in workbook.sheetnames:
                    raise ValueError(
                        f'Sheet {sheet!r} not found. Available: {", ".join(workbook.sheetnames)}'
                    )
                worksheet = workbook[sheet]
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([cell for cell in row])
        return rows, worksheet.title
    finally:
        workbook.close()


def read_rows(path, sheet=None):
    """Read any supported spreadsheet into (rows, sheet_name, source_file_basename)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in {'.csv', '.tsv'}:
        rows = read_csv_rows(path)
        sheet_name = os.path.basename(path)
    elif ext in {'.xlsx', '.xlsm'}:
        rows, sheet_name = read_xlsx_rows(path, sheet=sheet)
    elif ext == '.xls':
        raise ValueError('Legacy .xls is not supported. Save as .xlsx or CSV and upload again.')
    else:
        raise ValueError('Upload a .xlsx, .xlsm, .csv, or .tsv file.')
    return rows, sheet_name


def _require_openai_client():
    if OpenAI is None:
        raise RuntimeError('openai package is not installed.')
    api_key = os.environ.get('OPENAI_API_KEY', '')
    if not api_key:
        raise RuntimeError('OPENAI_API_KEY is not set.')
    return OpenAI(api_key=api_key)


def ai_guess_column_map(headers, sample_rows, canonical_fields, field_hints, model_name):
    """Ask an LLM to map headers -> canonical fields when the synonym dictionary can't.

    Returns {field: header_text}, restricted to real headers and real canonical
    fields — never trusts the model's output blindly. Raises on any failure
    (missing key, network error, bad JSON); callers should catch and treat
    that as "AI mapping unavailable" rather than let it break the request.
    """
    client = _require_openai_client()
    fields_desc = '\n'.join(f'- {field}: {field_hints.get(field, "")}' for field in canonical_fields)
    sample_lines = []
    for row in sample_rows[:AI_SAMPLE_ROWS]:
        cells = [_cell_text(c) for c in row]
        sample_lines.append(' | '.join(cells[: len(headers)]))
    prompt = (
        'You map spreadsheet column headers to a fixed set of database fields.\n'
        f'Headers (in order): {json.dumps(headers, ensure_ascii=False)}\n'
        f'Sample data rows:\n' + '\n'.join(sample_lines) + '\n\n'
        f'Target fields:\n{fields_desc}\n\n'
        'Return JSON only: {"mapping": {"<field>": "<exact header text from the list above>", ...}}. '
        'Only include a field if you are confident which header it is. '
        'Use the sample data to disambiguate (e.g. a column full of email-looking values is an email field). '
        'Do not invent header names that are not in the list.'
    )
    resp = client.chat.completions.create(
        model=model_name,
        response_format={'type': 'json_object'},
        messages=[{'role': 'user', 'content': prompt}],
        temperature=0,
    )
    content = resp.choices[0].message.content or '{}'
    data = json.loads(content)
    raw_mapping = data.get('mapping') or {}
    if not isinstance(raw_mapping, dict):
        return {}

    header_set = set(headers)
    safe_mapping = {}
    for field, header_name in raw_mapping.items():
        if field not in canonical_fields:
            continue
        if header_name not in header_set:
            continue
        safe_mapping[field] = headers.index(header_name)
    return safe_mapping
